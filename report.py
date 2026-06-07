"""Test report generation: per-step JSON results and heatmap Excel."""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from gepa.core.adapter import EvaluationBatch


def compute_classification_metrics(
    eval_batch: EvaluationBatch,
) -> dict[str, Any] | None:
    """Compute macro/micro precision, recall, f1 and accuracy from classification trajectories.

    Returns None if trajectories are missing or don't contain label data.
    """
    if not eval_batch.trajectories:
        return None

    pairs: list[tuple[str, str]] = []
    for traj in eval_batch.trajectories:
        for step in traj.get("steps_trace", []):
            expected = step.get("expected_label")
            actual = step.get("actual_label")
            if expected is None or actual is None:
                return None
            pairs.append((expected, actual or ""))

    if not pairs:
        return None

    labels = sorted({lbl for pair in pairs for lbl in pair})

    tp: dict[str, int] = {l: 0 for l in labels}
    fp: dict[str, int] = {l: 0 for l in labels}
    fn: dict[str, int] = {l: 0 for l in labels}

    for expected, actual in pairs:
        if actual == expected:
            tp[expected] += 1
        else:
            fn[expected] += 1
            if actual in fp:
                fp[actual] += 1

    # Per-class metrics
    per_class: dict[str, dict[str, float]] = {}
    for label in labels:
        p = tp[label] / (tp[label] + fp[label]) if (tp[label] + fp[label]) > 0 else 0.0
        r = tp[label] / (tp[label] + fn[label]) if (tp[label] + fn[label]) > 0 else 0.0
        f1 = 2 * p * r / (p + r) if (p + r) > 0 else 0.0
        per_class[label] = {"precision": round(p, 4), "recall": round(r, 4), "f1": round(f1, 4), "support": tp[label] + fn[label]}

    # Macro: average over classes with support > 0
    active = [v for v in per_class.values() if v["support"] > 0]
    n = len(active) or 1
    macro_precision = round(sum(v["precision"] for v in active) / n, 4)
    macro_recall = round(sum(v["recall"] for v in active) / n, 4)
    macro_f1 = round(sum(v["f1"] for v in active) / n, 4)

    # Micro: global TP / (TP + FP), TP / (TP + FN)
    total_tp = sum(tp.values())
    total_fp = sum(fp.values())
    total_fn = sum(fn.values())
    micro_precision = round(total_tp / (total_tp + total_fp), 4) if (total_tp + total_fp) > 0 else 0.0
    micro_recall = round(total_tp / (total_tp + total_fn), 4) if (total_tp + total_fn) > 0 else 0.0
    micro_f1 = round(2 * micro_precision * micro_recall / (micro_precision + micro_recall), 4) if (micro_precision + micro_recall) > 0 else 0.0

    accuracy = round(sum(1 for e, a in pairs if e == a) / len(pairs), 4)

    return {
        "macro_precision": macro_precision,
        "macro_recall": macro_recall,
        "macro_f1": macro_f1,
        "micro_precision": micro_precision,
        "micro_recall": micro_recall,
        "micro_f1": micro_f1,
        "accuracy": accuracy,
    }


def _normalize_scenario_id(scenario_id: str) -> str:
    """Strip '/modification_X' suffix so modifications aggregate into base scenario."""
    return scenario_id.split("/")[0].strip()


def build_test_results(
    eval_batch: EvaluationBatch,
    label: str = "",
) -> list[dict[str, Any]]:
    """Convert EvaluationBatch with trajectories into notebook-compatible results list."""
    if eval_batch.trajectories is None:
        return []

    results: list[dict[str, Any]] = []
    for trajectory in eval_batch.trajectories:
        for step_trace in trajectory["steps_trace"]:
            # UnblockCard/ViS traces use expected_message/actual_response;
            # KIB traces use expected_label/actual_label/actual_raw.
            result_entry: dict[str, Any] = {
                "scenario_id": trajectory["scenario_id"],
                "scenario_description": trajectory.get("scenario_description", ""),
                "dialog_id": trajectory["dialog_id"],
                "step": step_trace["step"],
                "status": "PASS" if step_trace["passed"] else "FAIL",
                "label": label,
            }
            if "expected_message" in step_trace:
                result_entry["agent_action_type"] = step_trace.get("action_type", "")
                result_entry["expected_agent_message"] = step_trace["expected_message"]
                result_entry["expected_pattern"] = step_trace["expected_pattern"]
                result_entry["agent_response"] = step_trace["actual_response"]
            else:
                result_entry["expected_label"] = step_trace.get("expected_label", "")
                result_entry["actual_label"] = step_trace.get("actual_label", "")
                result_entry["user_message"] = step_trace.get("user_message", "")
            results.append(result_entry)
    return results


def save_test_results_json(results: list[dict[str, Any]], path: Path) -> None:
    """Save per-step test results to JSON."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")


def generate_heatmap_excel(
    results: list[dict[str, Any]],
    output_path: Path,
    template_path: Path | None = None,
    header_row: int = 4,
    first_data_row: int = 5,
    scenario_id_col: int = 3,
    first_type_col: int = 5,
) -> None:
    """Generate heatmap Excel with scenario × action_type accuracy rates.

    If template_path is provided, fills an existing template.
    Otherwise creates a new workbook from scratch.
    """
    if not results:
        return

    # --- Normalize scenario IDs (strip /modification_X suffixes) ---
    # --- Aggregate results: scenario_id × action_type → accuracy ---
    counts: dict[str, dict[str, dict[str, int]]] = defaultdict(lambda: defaultdict(lambda: {"passed": 0, "total": 0}))
    for r in results:
        scenario_id = _normalize_scenario_id(r["scenario_id"])
        action_type = r.get("agent_action_type") or "Тип не указан"
        counts[scenario_id][action_type]["total"] += 1
        if r["status"] == "PASS":
            counts[scenario_id][action_type]["passed"] += 1

    rates: dict[str, dict[str, float]] = {
        sc: {
            at: d["passed"] / d["total"] if d["total"] > 0 else 0.0
            for at, d in types.items()
        }
        for sc, types in counts.items()
    }

    # Sorted unique action types → numeric IDs
    all_types = sorted({at for types in counts.values() for at in types})
    type_to_id = {t: str(i + 1) for i, t in enumerate(all_types)}

    # Build scenario_id → description mapping for standalone heatmaps.
    scenario_descs: dict[str, str] = {}
    for r in results:
        sid = _normalize_scenario_id(r["scenario_id"])
        if sid not in scenario_descs:
            scenario_descs[sid] = r.get("scenario_description", "")

    if template_path and template_path.exists():
        _fill_template(rates, type_to_id, template_path, output_path,
                       header_row, first_data_row, scenario_id_col, first_type_col)
    else:
        _create_standalone(rates, type_to_id, all_types, output_path, scenario_descs)


def _fill_template(
    rates: dict[str, dict[str, float]],
    type_to_id: dict[str, str],
    template_path: Path,
    output_path: Path,
    header_row: int,
    first_data_row: int,
    scenario_id_col: int,
    first_type_col: int,
) -> None:
    """Fill existing Excel template (same logic as notebook)."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Write type IDs to header row (template may have empty type columns).
    sorted_type_ids = sorted(type_to_id.values(), key=int)
    for i, tid in enumerate(sorted_type_ids):
        ws.cell(row=header_row, column=first_type_col + i, value=int(tid))

    # Build column map from header row
    col_map = {
        str(ws.cell(row=header_row, column=c).value): c
        for c in range(first_type_col, ws.max_column + 1)
        if ws.cell(row=header_row, column=c).value is not None
    }

    # Build row map from scenario ID column
    row_map = {
        ws.cell(row=r, column=scenario_id_col).value: r
        for r in range(first_data_row, ws.max_row + 1)
        if ws.cell(row=r, column=scenario_id_col).value is not None
    }

    fill_count = 0
    for scenario_id, type_rates in rates.items():
        target_row = row_map.get(scenario_id)
        if not target_row:
            continue
        for action_type, rate in type_rates.items():
            action_type_id = type_to_id.get(action_type)
            target_col = col_map.get(action_type_id)
            if not target_col:
                continue
            ws.cell(row=target_row, column=target_col).value = rate
            fill_count += 1

    # Add legend
    id_to_type = {v: k for k, v in type_to_id.items()}
    legend_start_row = ws.max_row + 3
    ws.cell(row=legend_start_row, column=scenario_id_col - 1, value="Расшифровка типов действий:")
    for type_id in sorted(id_to_type.keys(), key=int):
        ws.append({scenario_id_col - 1: type_id, first_type_col - 2: id_to_type[type_id]})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Heatmap (template): {output_path}  ({fill_count} cells filled)")


def _create_standalone(
    rates: dict[str, dict[str, float]],
    type_to_id: dict[str, str],
    all_types: list[str],
    output_path: Path,
    scenario_descs: dict[str, str] | None = None,
) -> None:
    """Create a standalone heatmap Excel matching the original ВиС format.

    Layout (1-indexed):
      Row 2: "Общая точность:" + formula
      Row 4: header (ID | Тип действия | Описание | type_1 | type_2 | ... | Доля верных)
      Row 5+: data rows per scenario_id
      After data: "Доля верных" summary row with AVERAGE formulas
      Below: legend (type ID → action type name)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Heatmap"

    sorted_scenarios = sorted(rates.keys())
    num_types = len(all_types)
    # Columns: B=ID, C=scenario_id, D=description, E..=types, last=Доля верных
    first_type_col = 5  # E
    last_type_col = first_type_col + num_types - 1
    avg_col = last_type_col + 1  # "Доля верных" column
    header_row = 4
    first_data_row = 5
    last_data_row = first_data_row + len(sorted_scenarios) - 1
    summary_row = last_data_row + 1

    # --- Row 2: overall accuracy ---
    ws.cell(row=2, column=2, value="Общая точность:")
    ws.cell(row=2, column=2).font = Font(bold=True, size=12)
    avg_col_letter = get_column_letter(avg_col)
    ws.cell(row=2, column=first_type_col,
            value=f"={avg_col_letter}{summary_row}")

    # --- Row 4: headers ---
    ws.cell(row=header_row, column=2, value="ID")
    ws.cell(row=header_row, column=3, value="Тип действия:")
    ws.cell(row=header_row, column=4, value="Описание")
    for i, action_type in enumerate(all_types):
        col = first_type_col + i
        ws.cell(row=header_row, column=col, value=int(type_to_id[action_type]))
    ws.cell(row=header_row, column=avg_col, value="Доля верных")
    for col in range(2, avg_col + 1):
        ws.cell(row=header_row, column=col).font = Font(bold=True)

    # --- Data rows ---
    # Collect scenario descriptions from rates metadata (stored in _scenario_descs).
    for idx, scenario_id in enumerate(sorted_scenarios):
        row = first_data_row + idx
        ws.cell(row=row, column=2, value=idx + 1)
        ws.cell(row=row, column=3, value=scenario_id)
        ws.cell(row=row, column=4, value=(scenario_descs or {}).get(scenario_id, ""))
        for i, action_type in enumerate(all_types):
            col = first_type_col + i
            rate = rates[scenario_id].get(action_type)
            if rate is not None:
                ws.cell(row=row, column=col, value=rate)
        # "Доля верных" per scenario: AVERAGE across type columns
        first_letter = get_column_letter(first_type_col)
        last_letter = get_column_letter(last_type_col)
        ws.cell(row=row, column=avg_col,
                value=f'=IFERROR(AVERAGE({first_letter}{row}:{last_letter}{row}),"-")')

    # --- Summary row: average per action type ---
    ws.cell(row=summary_row, column=3, value="Доля верных")
    ws.cell(row=summary_row, column=3).font = Font(bold=True)
    for i in range(num_types):
        col = first_type_col + i
        col_letter = get_column_letter(col)
        ws.cell(row=summary_row, column=col,
                value=f'=IFERROR(AVERAGE({col_letter}{first_data_row}:{col_letter}{last_data_row}),"-")')
    # Overall average
    first_letter = get_column_letter(first_type_col)
    last_letter = get_column_letter(last_type_col)
    ws.cell(row=summary_row, column=avg_col,
            value=f'=IFERROR(AVERAGE({first_letter}{first_data_row}:{last_letter}{last_data_row}),"-")')

    # --- Conditional formatting: green ≥0.8, yellow ≥0.5, red <0.5 ---
    data_range = f"{get_column_letter(first_type_col)}{first_data_row}:{get_column_letter(avg_col)}{summary_row}"
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(data_range, CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=green_fill))
    ws.conditional_formatting.add(data_range, CellIsRule(operator="between", formula=["0.5", "0.7999"], fill=yellow_fill))
    ws.conditional_formatting.add(data_range, CellIsRule(operator="lessThan", formula=["0.5"], fill=red_fill))

    # --- Legend ---
    legend_row = summary_row + 3
    ws.cell(row=legend_row, column=2, value="Расшифровка типов действий:")
    ws.cell(row=legend_row, column=2).font = Font(bold=True)
    id_to_type = {v: k for k, v in type_to_id.items()}
    for type_id in sorted(id_to_type.keys(), key=int):
        legend_row += 1
        ws.cell(row=legend_row, column=2, value=type_id)
        ws.cell(row=legend_row, column=3, value=id_to_type[type_id])

    # --- Column widths ---
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 60
    for i in range(num_types):
        ws.column_dimensions[get_column_letter(first_type_col + i)].width = 8
    ws.column_dimensions[get_column_letter(avg_col)].width = 14

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Heatmap (standalone): {output_path}")
