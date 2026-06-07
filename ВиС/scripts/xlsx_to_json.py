"""Convert ВиС Excel run results to the pipeline JSON dataset format.

Usage:
    python xlsx_to_json.py <input.xlsx> <output.json>

Example:
    python xlsx_to_json.py \
        "../Прогоны/dataset_ViS_v12_prompt_ViS_v14_GigaChat-2-Pro_run_1.xlsx" \
        "../datasets/from_excel.json"
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

import openpyxl


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Convert ВиС Excel to JSON dataset.")
    p.add_argument("input", help="Path to .xlsx file")
    p.add_argument("output", help="Path to output .json file")
    return p.parse_args()


def _build_steps_from_rows(rows: list[tuple], col: dict[str, int]) -> list[dict]:
    """Build the steps list from sorted Excel rows belonging to one test case."""
    steps = []
    for row in rows:
        step_user = int(row[col["step_user"]])
        step_agent = int(row[col["step_agent"]])

        # Reconstruct intermediate agent steps from dialog_history if needed.
        # For the first row of a multi-step dialog (step_user > 1),
        # earlier steps are embedded in dialog_history_text; here we only
        # record the current user→agent exchange since the pipeline adapter
        # replays earlier turns from the dialog_history field itself.

        # User step
        steps.append({
            "step": step_user,
            "role": "user",
            "message": str(row[col["user_message"]]),
        })

        # Agent step
        agent_step: dict = {
            "step": step_agent,
            "role": "agent",
            "type": str(row[col["agent_message_type"]]),
            "message": str(row[col["agent_message_etalon"]]),
        }
        pattern = row[col["expected_pattern"]]
        if pattern:
            agent_step["additional_check"] = str(pattern)
        steps.append(agent_step)

    return steps


def _parse_dialog_history(history_text: str | None) -> list[dict]:
    """Parse dialog_history_text into preceding steps.

    Format: "[Клиент] - msg1 \\n [Классификатор] - msg2 \\n ..."
    Returns list of step dicts with ascending step numbers.
    """
    if not history_text:
        return []

    steps = []
    step_num = 0
    for line in history_text.strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        if line.startswith("[Клиент]"):
            step_num += 1
            msg = line.replace("[Клиент]", "").lstrip(" -").strip()
            steps.append({"step": step_num, "role": "user", "message": msg})
        elif line.startswith("[Классификатор]"):
            step_num += 1
            msg = line.replace("[Классификатор]", "").lstrip(" -").strip()
            # We mark intermediate agent steps as "Сообщение человеку" with a
            # permissive pattern — the pipeline only scores the final agent turn.
            steps.append({
                "step": step_num,
                "role": "agent",
                "type": "Сообщение человеку",
                "message": msg,
                "additional_check": "^КЛИЕНТ(?!.*\\bFAQ\\b)(?!.*\\bОПЕРАТОР\\b)(?!.*\\b\\d+\\.\\s*[А-Я]).*",
            })
    return steps


def xlsx_to_dataset(path: str | Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    # Each unique test case is identified by the combination of
    # (scenario_id, dialog_id, dialog_history_text) to handle:
    # - Same dialog_id used in different scenarios
    # - Same dialog_id + scenario with different branches (different history)
    cases: dict[tuple, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        scenario_id = row[col["scenario_id"]]
        scenario_desc = row[col["scenario_description"]]
        dialog_id = row[col["dialog_id"]]
        history = row[col["dialog_history_text"]] or ""

        key = (scenario_id, dialog_id, history)
        if key not in cases:
            cases[key] = {
                "scenario_id": scenario_id,
                "scenario_description": scenario_desc,
                "dialog_id": dialog_id,
                "history": history,
                "rows": [],
            }
        cases[key]["rows"].append(row)

    dataset = []
    for key, info in cases.items():
        # Sort rows by step number
        sorted_rows = sorted(info["rows"], key=lambda r: (r[col["step_user"]], r[col["step_agent"]]))

        first_step_user = int(sorted_rows[0][col["step_user"]])

        if first_step_user > 1:
            # Multi-step dialog: reconstruct earlier steps from dialog_history
            preceding = _parse_dialog_history(info["history"])
            current = _build_steps_from_rows(sorted_rows, col)
            steps = preceding + current
        else:
            steps = _build_steps_from_rows(sorted_rows, col)

        dataset.append({
            "scenario_id": info["scenario_id"],
            "scenario_description": info["scenario_description"],
            "dialog_id": info["dialog_id"],
            "steps": steps,
        })

    wb.close()
    return dataset


def main() -> None:
    args = parse_args()
    dataset = xlsx_to_dataset(args.input)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Converted {len(dataset)} dialogs to {out}")


if __name__ == "__main__":
    main()
